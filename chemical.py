import requests
from bs4 import BeautifulSoup
import pandas as pd
from google_trans_new import google_translator
import openpyxl 

translator = google_translator()
#爬虫去爬的网站
url = 'https://labchem-wako.fujifilm.com/jp/product/result/product.html?fw='
url_2='https://www.dinochem.com/product/'
url_3='https://m.chemicalbook.com/Search_JP.aspx?keyword='
#存放地址
sheet_address='C:/Users/Kai Isobe/Desktop/sheet/'
#要读的表名称
sheet_name='210219q.xlsx'

#sheet = pd.read_excel(sheet_address+sheet_name)

wb = openpyxl.load_workbook(sheet_address+sheet_name)
sheet_ranges = wb['Sheet1']

html = requests.get(url_3+'117241-32-4').text
soup = BeautifulSoup(html,'html.parser')
print(soup.select('ul')[1].contents[5].contents[0].string)
print(soup.select('ul')[1].contents[7].contents[0].string)
print(soup.select('ul')[1].contents[5].contents[2])
print(soup.select('ul')[1].contents[7].contents[2])

#全角转半角
def dbc_to_sbc(ustring):

 rstring =''
 for uchar in ustring:
    inside_code = ord(uchar)
    if inside_code == 0x3000:
      inside_code = 0x0020
    else:
      inside_code -= 0xfee0
    if not (0x0021 <= inside_code and inside_code <= 0x7e):
      rstring += uchar
      continue
    rstring += chr(inside_code)
 return rstring

#查找英文和日文名称

def cas_name():
  n=0
  url_merge=''
  for cas_no in sheet_ranges['D']:
  #for cas_no in sheet['cas_rn'].values:
    cas_no=cas_no.value
    if cas_no=='cas_rn':continue
    n+=1
    print(n)
    
    url_merge=url+cas_no
    html = requests.get(url_merge).text
    soup = BeautifulSoup(html,'html.parser')
    try:
      ja_name = soup.em.contents[0].rstrip()
      en_name = soup.em.q.string
    except:
      ja_name=''
      en_name=''
    
    if ja_name == en_name: ja_name=''
    if not en_name or not ja_name:
      try:
        print('换个网站找')
        url_merge=url_3+cas_no
        html = requests.get(url_merge).text
        soup = BeautifulSoup(html,'html.parser')
        if not en_name:en_name = soup.select('ul')[1].contents[5].contents[2]
        if not ja_name:ja_name = soup.select('ul')[1].contents[7].contents[2]
      except:
        print('ss')
    if not en_name:
      try:
        print('再换个网站找')
        url_merge=url_2+cas_no+'.html'
        html=requests.get(url_merge).text
        soup = BeautifulSoup(html,'html.parser')
        en_name=soup.b.string

      except:
        en_name=''
    if en_name and not ja_name:
      print('开始翻译')
      ja_name = translator.translate(en_name,lang_tgt='ja').replace(' ','')
      ja_name = dbc_to_sbc(ja_name)
    
    yield [cas_no,ja_name,en_name,url_merge]


#执行程序
def main():
  row=0
  for name_ in cas_name():
    # cas_no=name_[0]
    # sheet.iloc[row,4]=name_[1]
    # sheet.iloc[row,5]=name_[2]
    # sheet.iloc[row,11]=name_[3]
    sheet_ranges.cell(row+2,5).value=name_[1]
    sheet_ranges.cell(row+2,6).value=name_[2]
    sheet_ranges.cell(row+2,12).value=name_[3]
    row+=1
  #sheet.to_excel('C:/Users/Kai Isobe/Desktop/sheet/2.xlsx', index=False, header=True)
  wb.save(sheet_address+'newsheet/'+'new_'+sheet_name)
#main()
  

























